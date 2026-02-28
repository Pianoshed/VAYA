import csv
import json
import os
import random
from datetime import datetime, timedelta

import pandas as pd
import requests
from sqlalchemy import func
from dotenv import load_dotenv
from flask import Flask, render_template, request, redirect, url_for, session, jsonify, make_response
from flask import send_file
from flask_sqlalchemy import SQLAlchemy
from openpyxl import load_workbook
from werkzeug.security import check_password_hash, generate_password_hash

# Load environment variables
load_dotenv()

app = Flask(__name__)

# Database configurations
app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('SQLALCHEMY_DATABASE_URI', 'sqlite:///default.db')
app.config['SQLALCHEMY_BINDS'] = {
    'mood_tracker': os.getenv('SQLALCHEMY_BINDS_MOOD_TRACKER', 'sqlite:///mood_tracker.db'),
    'submit_assessment': os.getenv('SQLALCHEMY_BINDS_SUBMIT_ASSESSMENT', 'sqlite:///submit-assessment.db'),
}
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# Securely load secret key
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'Akingbesote Babajide Created This Site 08134812419')

# Placeholder for tracking homepage visits
homepage_visits = []

JOURNAL_FILE = "journal_entries.json"

# Upload folder setup
UPLOAD_FOLDER = "static/uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

CSV_FILE = os.path.join(UPLOAD_FOLDER, "responses.csv")
EXCEL_FILE = os.path.join(UPLOAD_FOLDER, "responses.xlsx")

# Game progress data file
GAME_PROGRESS_FILE = os.path.join(UPLOAD_FOLDER, 'game_progress.json')

# Admin credentials (hashed password)
ADMIN_USERNAME = os.getenv('ADMIN_USERNAME', 'admin')
ADMIN_PASSWORD_HASH = generate_password_hash(os.getenv('ADMIN_PASSWORD', 'admin123'))

# API Keys
IPGEO_API_KEY = os.getenv('IPGEO_API_KEY')

# Initialize game progress file if it doesn't exist
if not os.path.exists(GAME_PROGRESS_FILE):
    with open(GAME_PROGRESS_FILE, 'w') as f:
        json.dump([], f)


def load_game_data():
    """Load game progress data from file"""
    try:
        with open(GAME_PROGRESS_FILE, 'r') as f:
            return json.load(f)
    except:
        return []


def save_game_data(data):
    """Save game progress data to file"""
    with open(GAME_PROGRESS_FILE, 'w') as f:
        json.dump(data, f, indent=2)


# Function to load journal entries
def load_journal_entries():
    if os.path.exists(JOURNAL_FILE):
        with open(JOURNAL_FILE, "r") as file:
            return json.load(file)
    return []


# Function to save a new journal entry
def save_journal_entry(entry):
    entries = load_journal_entries()
    entries.append(entry)
    with open(JOURNAL_FILE, "w") as file:
        json.dump(entries, file, indent=4)


# ── MODELS ──────────────────────────────────────────────────────────────────

class JournalEntry(db.Model):
    __bind_key__ = 'mood_tracker'
    id = db.Column(db.Integer, primary_key=True)
    mood = db.Column(db.String(50), nullable=False)
    entry = db.Column(db.Text, nullable=False)
    activity = db.Column(db.String(255))
    note = db.Column(db.Text)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)


class Registration(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100))
    age = db.Column(db.Integer)
    email = db.Column(db.String(100))
    phone = db.Column(db.String(20))
    service = db.Column(db.String(100))

    def __repr__(self):
        return f'<Registration {self.name}>'


class SubmitAssessment(db.Model):
    __bind_key__ = 'submit_assessment'
    id = db.Column(db.Integer, primary_key=True)
    timestamp = db.Column(db.String(100), default=lambda: datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    email = db.Column(db.String(100))
    phone = db.Column(db.String(20))
    age = db.Column(db.Integer)
    sex = db.Column(db.String(20))
    score = db.Column(db.Integer)
    q1 = db.Column(db.Integer);  q2 = db.Column(db.Integer);  q3 = db.Column(db.Integer)
    q4 = db.Column(db.Integer);  q5 = db.Column(db.Integer);  q6 = db.Column(db.Integer)
    q7 = db.Column(db.Integer);  q8 = db.Column(db.Integer);  q9 = db.Column(db.Integer)
    q10 = db.Column(db.Integer); q11 = db.Column(db.Integer); q12 = db.Column(db.Integer)
    q13 = db.Column(db.Integer); q14 = db.Column(db.Integer); q15 = db.Column(db.Integer)
    q16 = db.Column(db.Integer); q17 = db.Column(db.Integer); q18 = db.Column(db.Integer)
    q19 = db.Column(db.Integer); q20 = db.Column(db.Integer)


class Feedback(db.Model):
    __bind_key__ = 'submit_assessment'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(100), nullable=False)
    message = db.Column(db.Text, nullable=False)
    submitted_at = db.Column(db.DateTime, default=datetime.utcnow)


class PageVisit(db.Model):
    """Tracks every page visit with route, IP, location and timestamp."""
    id = db.Column(db.Integer, primary_key=True)
    ip_address = db.Column(db.String(100))
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)
    location = db.Column(db.String(255))
    page = db.Column(db.String(100), default='/')          # which route was visited
    user_agent = db.Column(db.String(255))                 # browser / device info


with app.app_context():
    db.create_all()


# ── HELPER: record a page visit ──────────────────────────────────────────────

def record_page_visit(page: str):
    """Call this at the top of any route you want to monitor."""
    ip_address = request.remote_addr
    user_agent = request.headers.get('User-Agent', 'Unknown')
    location = get_location(ip_address)

    visit = PageVisit(
        ip_address=ip_address,
        location=location,
        page=page,
        user_agent=user_agent,
    )
    db.session.add(visit)
    db.session.commit()

    # Also keep the in-memory log for the / route
    if page == '/':
        homepage_visits.append({
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "location": location
        })


# Random Challenges
challenges = [
    "🌞 Take a 5-minute walk outside!",
    "🎶 Listen to your favorite song and dance!",
    "📖 Read a page from an inspiring book.",
    "🌿 Try deep breathing for 1 minute.",
    "💌 Send a kind message to a friend.",
    "🎨 Draw or doodle something fun!"
]

# Random Affirmations
affirmations = [
    "💖 You are enough just as you are.",
    "🌟 You have the strength to overcome any challenge.",
    "😊 You bring positivity into the world.",
    "💙 You deserve love, happiness, and peace.",
    "🌿 Take a deep breath. You are doing great.",
    "🔥 You are capable of amazing things."
]

# ── SCORE MAPPINGS (single source of truth) ──────────────────────────────────
SCORE_MAPPINGS = [
    {1: 1, 2: 2, 3: 3, 4: 4, 5: 5},   # Q1  POSITIVE 5-opt
    {1: 1, 2: 2, 3: 3, 4: 4, 5: 5},   # Q2  POSITIVE 5-opt
    {1: 1, 2: 2, 3: 3, 4: 4, 5: 5},   # Q3  POSITIVE 5-opt
    {1: 1, 2: 2, 3: 3, 4: 4, 5: 5},   # Q4  POSITIVE 5-opt
    {1: 1, 2: 2, 3: 3, 4: 4, 5: 5},   # Q5  POSITIVE 5-opt
    {1: 4, 2: 3, 3: 2, 4: 1},          # Q6  NEGATIVE 4-opt
    {1: 4, 2: 3, 3: 2, 4: 1},          # Q7  NEGATIVE 4-opt
    {1: 5, 2: 4, 3: 3, 4: 2, 5: 1},   # Q8  NEGATIVE 5-opt
    {1: 5, 2: 4, 3: 3, 4: 2, 5: 1},   # Q9  NEGATIVE 5-opt
    {1: 5, 2: 4, 3: 3, 4: 2, 5: 1},   # Q10 NEGATIVE 5-opt
    {1: 1, 2: 2, 3: 3, 4: 4},          # Q11 POSITIVE 4-opt
    {1: 4, 2: 3, 3: 2, 4: 1},          # Q12 NEGATIVE CRITICAL 4-opt
    {1: 4, 2: 3, 3: 2, 4: 1},          # Q13 NEGATIVE CRITICAL 4-opt
    {1: 4, 2: 3, 3: 2, 4: 1},          # Q14 NEGATIVE 4-opt
    {1: 1, 2: 2, 3: 3, 4: 4},          # Q15 POSITIVE 4-opt
    {1: 1, 2: 2, 3: 3, 4: 4},          # Q16 POSITIVE 4-opt
    {1: 1, 2: 2, 3: 3, 4: 4},          # Q17 (higher = better context) 4-opt
    {1: 1, 2: 2, 3: 3, 4: 4},          # Q18 POSITIVE 4-opt
    {1: 1, 2: 2, 3: 3, 4: 4},          # Q19 POSITIVE 4-opt
    {1: 1, 2: 2, 3: 3, 4: 4},          # Q20 POSITIVE 4-opt
]


# ── ANALYTICS HELPER ─────────────────────────────────────────────────────────

def calculate_analytics():
    """Calculate comprehensive analytics for the admin dashboard."""
    assessments = SubmitAssessment.query.all()

    empty = {
        'total_assessments': 0,
        'avg_score': 0,
        'median_score': 0,
        'min_score': 0,
        'max_score': 0,
        'score_distribution': {
            '0-17 (Critical)': 0, '18-34 (High Risk)': 0,
            '35-51 (Moderate)': 0, '52-69 (Good)': 0, '70-88 (Excellent)': 0
        },
        'gender_breakdown': {'male': 0, 'female': 0, 'other': 0},
        'age_distribution': {'10-15': 0, '16-20': 0, '21-25': 0, '26-30': 0, '31-35': 0},
        'question_averages': {},
        'recent_trend': [],
        'risk_categories': {'critical': 0, 'high_risk': 0, 'moderate': 0, 'good': 0, 'excellent': 0},
        'completion_rate_last_7_days': 0,
        'top_concerns': []
    }

    if not assessments:
        return empty

    # Build scores list
    scores = []
    for a in assessments:
        if hasattr(a, 'score') and a.score is not None:
            scores.append(a.score)
        else:
            total = 0
            for i in range(1, 21):
                q_val = getattr(a, f'q{i}', None)
                if q_val is not None:
                    total += SCORE_MAPPINGS[i - 1].get(q_val, 0)
            scores.append(total)

    total_count = len(assessments)
    avg_score = sum(scores) / len(scores) if scores else 0
    median_score = sorted(scores)[len(scores) // 2] if scores else 0
    min_score = min(scores) if scores else 0
    max_score = max(scores) if scores else 0

    score_ranges = {
        '0-17 (Critical)': 0, '18-34 (High Risk)': 0,
        '35-51 (Moderate)': 0, '52-69 (Good)': 0, '70-88 (Excellent)': 0
    }
    for score in scores:
        if score <= 17:
            score_ranges['0-17 (Critical)'] += 1
        elif score <= 34:
            score_ranges['18-34 (High Risk)'] += 1
        elif score <= 51:
            score_ranges['35-51 (Moderate)'] += 1
        elif score <= 69:
            score_ranges['52-69 (Good)'] += 1
        else:
            score_ranges['70-88 (Excellent)'] += 1

    gender_breakdown = {
        'male': SubmitAssessment.query.filter_by(sex='male').count(),
        'female': SubmitAssessment.query.filter_by(sex='female').count(),
        'other': SubmitAssessment.query.filter_by(sex='other').count()
    }

    ages = [a.age for a in assessments if a.age is not None]
    age_distribution = {
        '10-15': sum(1 for age in ages if 10 <= age <= 15),
        '16-20': sum(1 for age in ages if 16 <= age <= 20),
        '21-25': sum(1 for age in ages if 21 <= age <= 25),
        '26-30': sum(1 for age in ages if 26 <= age <= 30),
        '31-35': sum(1 for age in ages if 31 <= age <= 35)
    }

    question_labels = {
        'q1': 'Cheerful and good spirits', 'q2': 'Calm and relaxed',
        'q3': 'Active and vigorous', 'q4': 'Wake up fresh and rested',
        'q5': 'Interested in daily life', 'q6': 'Little interest/pleasure',
        'q7': 'Down/depressed/hopeless', 'q8': 'Hard to concentrate',
        'q9': 'Nervous/anxious', 'q10': 'Emotions out of control',
        'q11': 'Safe at home', 'q12': 'Touched inappropriately',
        'q13': 'Recently hurt', 'q14': 'Pressured to do things',
        'q15': 'Can say No', 'q16': 'Sense of belonging',
        'q17': 'Feel lonely', 'q18': 'Optimistic about future',
        'q19': 'Can handle problems', 'q20': 'Know where to get help'
    }

    question_averages = {}
    for i in range(1, 21):
        q_field = f'q{i}'
        responses = [getattr(a, q_field) for a in assessments if getattr(a, q_field) is not None]
        if responses:
            avg = sum(responses) / len(responses)
            question_averages[q_field] = {
                'label': question_labels.get(q_field, f'Question {i}'),
                'average': round(avg, 2),
                'total_responses': len(responses)
            }

    thirty_days_ago = datetime.now() - timedelta(days=30)
    recent_assessments = []
    for a in assessments:
        try:
            if datetime.strptime(a.timestamp, "%Y-%m-%d %H:%M:%S") >= thirty_days_ago:
                recent_assessments.append(a)
        except Exception:
            pass

    recent_trend = []
    for i in range(30):
        date = (datetime.now() - timedelta(days=29 - i)).strftime("%Y-%m-%d")
        count = sum(1 for a in recent_assessments if a.timestamp.startswith(date))
        recent_trend.append({'date': date, 'count': count})

    risk_categories = {
        'critical': sum(1 for s in scores if s <= 17),
        'high_risk': sum(1 for s in scores if 18 <= s <= 34),
        'moderate': sum(1 for s in scores if 35 <= s <= 51),
        'good': sum(1 for s in scores if 52 <= s <= 69),
        'excellent': sum(1 for s in scores if s >= 70)
    }

    seven_days_ago = datetime.now() - timedelta(days=7)
    recent_week = []
    for a in assessments:
        try:
            if datetime.strptime(a.timestamp, "%Y-%m-%d %H:%M:%S") >= seven_days_ago:
                recent_week.append(a)
        except Exception:
            pass
    completion_rate_last_7_days = len(recent_week)

    concern_questions = []
    for q_num in [6, 7, 8, 9, 10, 13, 14, 17]:
        q_field = f'q{q_num}'
        responses = [getattr(a, q_field) for a in assessments if getattr(a, q_field) is not None]
        if responses:
            avg = sum(responses) / len(responses)
            concern_questions.append({
                'question': question_labels.get(q_field, f'Question {q_num}'),
                'average': round(avg, 2),
                'severity': 'High' if avg >= 3.5 else 'Moderate' if avg >= 2.5 else 'Low'
            })

    top_concerns = sorted(concern_questions, key=lambda x: x['average'], reverse=True)[:5]

    return {
        'total_assessments': total_count,
        'avg_score': round(avg_score, 2),
        'median_score': median_score,
        'min_score': min_score,
        'max_score': max_score,
        'score_distribution': score_ranges,
        'gender_breakdown': gender_breakdown,
        'age_distribution': age_distribution,
        'question_averages': question_averages,
        'recent_trend': recent_trend,
        'risk_categories': risk_categories,
        'completion_rate_last_7_days': completion_rate_last_7_days,
        'top_concerns': top_concerns
    }


def calculate_page_analytics():
    """Aggregate page visit data for the admin dashboard."""
    all_visits = PageVisit.query.order_by(PageVisit.timestamp.desc()).all()
    total_visits = len(all_visits)
    unique_ips = db.session.query(func.count(func.distinct(PageVisit.ip_address))).scalar() or 0

    # Per-page counts
    page_counts = {}
    for v in all_visits:
        page = v.page or '/'
        page_counts[page] = page_counts.get(page, 0) + 1

    # Top pages sorted
    top_pages = sorted(page_counts.items(), key=lambda x: x[1], reverse=True)

    # Daily visits for last 30 days
    thirty_days_ago = datetime.now() - timedelta(days=30)
    recent = [v for v in all_visits if v.timestamp and v.timestamp >= thirty_days_ago]
    daily_visits = {}
    for v in recent:
        day = v.timestamp.strftime("%Y-%m-%d")
        daily_visits[day] = daily_visits.get(day, 0) + 1

    daily_trend = []
    for i in range(30):
        date = (datetime.now() - timedelta(days=29 - i)).strftime("%Y-%m-%d")
        daily_trend.append({'date': date, 'count': daily_visits.get(date, 0)})

    # Visits last 7 days
    seven_days_ago = datetime.now() - timedelta(days=7)
    visits_last_7 = sum(1 for v in all_visits if v.timestamp and v.timestamp >= seven_days_ago)

    # Top locations
    location_counts = {}
    for v in all_visits:
        loc = v.location or 'Unknown'
        location_counts[loc] = location_counts.get(loc, 0) + 1
    top_locations = sorted(location_counts.items(), key=lambda x: x[1], reverse=True)[:10]

    # Recent visits (last 20)
    recent_visits = all_visits[:20]

    return {
        'total_visits': total_visits,
        'unique_visitors': unique_ips,
        'visits_last_7_days': visits_last_7,
        'top_pages': top_pages,
        'daily_trend': daily_trend,
        'top_locations': top_locations,
        'recent_visits': recent_visits,
        'page_counts': page_counts,
    }


# ── COOKIE ROUTES ─────────────────────────────────────────────────────────────

@app.route('/set-cookie', methods=['POST'])
def set_cookie():
    data = request.json
    cookie_name = data.get('name')
    cookie_value = data.get('value')
    days = data.get('days', 30)
    if not cookie_name or not cookie_value:
        return jsonify({"error": "Missing cookie name or value"}), 400
    response = make_response(jsonify({"message": f"Cookie {cookie_name} set"}))
    response.set_cookie(cookie_name, cookie_value, max_age=days * 24 * 60 * 60)
    return response


@app.route('/get-cookie', methods=['GET'])
def get_cookie():
    cookie_name = request.args.get('name')
    if not cookie_name:
        return jsonify({"error": "Cookie name is required"}), 400
    cookie_value = request.cookies.get(cookie_name)
    if cookie_value:
        return jsonify({cookie_name: cookie_value})
    return jsonify({"error": "Cookie not found"}), 404


# ── MAIN ROUTES ───────────────────────────────────────────────────────────────

@app.route('/')
def index():
    record_page_visit('/')

    ip_address = request.remote_addr
    location = get_location(ip_address)

    visit_count = request.cookies.get("visit_count", 0)
    visit_count = int(visit_count) + 1

    resp = make_response(render_template("index.html", location=location, visit_count=visit_count))
    resp.set_cookie("visit_count", str(visit_count), max_age=365 * 24 * 60 * 60, httponly=True, samesite="Lax")
    return resp


@app.route('/mood_tracker')
def mood_tracker():
    record_page_visit('/mood_tracker')
    return render_template('mood_tracker.html')


@app.route('/mental')
def mental():
    record_page_visit('/mental')
    return render_template('mental.html')


@app.route('/Resources')
def Resources():
    record_page_visit('/Resources')
    return render_template('Resources.html')


@app.route('/sel_activities')
def sel_activities():
    record_page_visit('/sel_activities')
    return render_template('sel-activities.html')


@app.route('/register', endpoint='register')
def register():
    record_page_visit('/register')
    return render_template('register.html')


@app.route('/runnerapp')
def runner_app():
    record_page_visit('/runnerapp')
    return render_template('runnerapp.html')


@app.route('/visit')
def visit():
    visit_count = request.cookies.get('visit_count', 0)
    location = request.cookies.get('location', "Unknown")
    return render_template('visit.html', visit_count=visit_count, homepage_visits=homepage_visits)


@app.route('/confirmation')
def confirmation():
    latest_submission = Registration.query.order_by(Registration.id.desc()).first()
    if latest_submission:
        return render_template(
            'confirmation.html',
            name=latest_submission.name,
            age=latest_submission.age,
            email=latest_submission.email,
            phone=latest_submission.phone,
            service=latest_submission.service,
        )
    return "No submission found."


# ── API ROUTES ────────────────────────────────────────────────────────────────

@app.route('/api/affirmation', methods=['GET'])
def get_affirmation():
    today = datetime.now().strftime("%Y-%m-%d")
    random.seed(today + "affirmation")
    daily_affirmation = random.choice(affirmations)
    random.seed()
    return jsonify({"affirmation": daily_affirmation})


@app.route('/api/challenge', methods=['GET'])
def get_challenge():
    today = datetime.now().strftime("%Y-%m-%d")
    random.seed(today + "challenge")
    daily_challenge = random.choice(challenges)
    random.seed()
    return jsonify({"challenge": daily_challenge})


@app.route('/api/journal', methods=['GET'])
def get_journals():
    journals = JournalEntry.query.all()
    journal_list = [{"id": j.id, "mood": j.mood, "entry": j.entry} for j in journals]
    return jsonify(journal_list)


@app.route('/save_journal_entry', methods=['POST'])
def save_journal_entry_route():
    data = request.get_json()
    if not data or 'mood' not in data or 'entry' not in data:
        return jsonify({'error': 'Mood and entry are required'}), 400
    new_entry = JournalEntry(
        mood=data['mood'],
        entry=data['entry'],
        activity=data.get('activity', None),
        note=data.get('note', None),
        timestamp=data.get('timestamp', datetime.utcnow())
    )
    db.session.add(new_entry)
    db.session.commit()
    return jsonify({'message': 'Journal entry saved successfully'}), 201


@app.route('/api/get_journal')
def get_journal():
    entries = load_journal_entries()
    return jsonify(entries)


# ── FORM SUBMISSION ROUTES ────────────────────────────────────────────────────

def calculate_age(dob):
    try:
        birth_date = datetime.strptime(dob, "%Y-%m-%d")
        today = datetime.today()
        return today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))
    except ValueError:
        return None


@app.route('/submit', methods=['POST'])
def submit():
    required_fields = ['name', 'age', 'email', 'phone', 'service']
    for field in required_fields:
        if field not in request.form:
            return f"Missing field: {field}", 400

    name = request.form['name']
    age = calculate_age(request.form['age'])
    email = request.form['email']
    phone = request.form['phone']
    service = request.form['service']

    if not age:
        return "Invalid date format for age", 400

    new_registration = Registration(name=name, age=age, email=email, phone=phone, service=service)
    db.session.add(new_registration)
    db.session.commit()

    data = {key: [value] for key, value in request.form.items()}
    df = pd.DataFrame(data)

    if os.path.exists(EXCEL_FILE):
        with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            wb = load_workbook(EXCEL_FILE)
            sheet = wb.active
            start_row = sheet.max_row + 1
            df.to_excel(writer, index=False, header=False, startrow=start_row)
    else:
        df.to_excel(EXCEL_FILE, index=False)

    return redirect(url_for('confirmation'))


@app.route('/submit-assessment', methods=['POST'])
def submit_assessment():
    total_score = 0
    unanswered_questions = 0

    email = request.form.get("email")
    phone = request.form.get("phone")
    age   = request.form.get("age")
    sex   = request.form.get("sex")

    answers_dict = {}
    for i in range(1, 21):
        val = request.form.get(f'q{i}')
        if val:
            try:
                ans_int = int(val)
                mapping = SCORE_MAPPINGS[i - 1]
                if ans_int in mapping:
                    total_score += mapping[ans_int]
                    answers_dict[f'q{i}'] = ans_int
                else:
                    unanswered_questions += 1
            except (ValueError, KeyError):
                unanswered_questions += 1
        else:
            unanswered_questions += 1

    if unanswered_questions > 0:
        return render_template(
            'error.html',
            message=f"Please answer all questions. You missed {unanswered_questions}."
        )

    try:
        new_assessment = SubmitAssessment(
            email=email,
            phone=phone,
            age=age,
            sex=sex,
            score=total_score,
            **answers_dict
        )
        db.session.add(new_assessment)
        db.session.commit()
    except Exception as e:
        print(f"Database Error: {e}")

    try:
        data_for_excel = {
            "timestamp": [new_assessment.timestamp],
            "email": [email],
            "phone": [phone],
            "age": [age],
            "sex": [sex],
            "score": [total_score]
        }
        for i in range(1, 21):
            data_for_excel[f'q{i}'] = [answers_dict.get(f'q{i}', '')]

        df = pd.DataFrame(data_for_excel)

        if os.path.exists(EXCEL_FILE):
            with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                start_row = load_workbook(EXCEL_FILE).active.max_row
                df.to_excel(writer, index=False, header=False, startrow=start_row)
        else:
            df.to_excel(EXCEL_FILE, index=False)
    except Exception as e:
        print(f"Excel Error: {e}")

    return redirect(url_for('assessment_results', score=total_score))


@app.route('/assessment_results')
def assessment_results():
    score = request.args.get('score', 0, type=int)
    return render_template('assessment_results.html', score=score)


# ── ADMIN ROUTES ──────────────────────────────────────────────────────────────

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    error_message = None
    if request.method == 'POST':
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        if username == ADMIN_USERNAME and check_password_hash(ADMIN_PASSWORD_HASH, password):
            session["is_admin"] = True
            return redirect(url_for("admin_panel"))
        else:
            error_message = "Invalid username or password!"
    return render_template("admin_login.html", error=error_message)


@app.route('/admin/panel')
def admin_panel():
    if not session.get("is_admin"):
        return redirect(url_for("admin_login"))

    assessments = SubmitAssessment.query.order_by(SubmitAssessment.id.desc()).all()
    analytics = calculate_analytics()
    page_analytics = calculate_page_analytics()
    registrations = Registration.query.order_by(Registration.id.desc()).all()
    files = os.listdir(UPLOAD_FOLDER) if os.path.exists(UPLOAD_FOLDER) else []

    total_visits = PageVisit.query.count()
    unique_visitors = db.session.query(func.count(func.distinct(PageVisit.ip_address))).scalar() or 0

    game_data = load_game_data()
    game_stats = {
        'total_sessions': len(game_data),
        'total_glimmers': sum(entry.get('glimmers', 0) for entry in game_data),
        'average_glimmers': round(sum(entry.get('glimmers', 0) for entry in game_data) / len(game_data), 2) if game_data else 0,
        'highest_score': max((entry.get('glimmers', 0) for entry in game_data), default=0),
        'latest_target': game_data[-1].get('target', 10) if game_data else 10
    }

    return render_template(
        "admin_panel.html",
        assessments=assessments,
        analytics=analytics,
        page_analytics=page_analytics,
        registrations=registrations,
        files=files,
        total_visits=total_visits,
        unique_visitors=unique_visitors,
        game_stats=game_stats
    )


@app.route('/admin/registrations')
def admin_registrations():
    if not session.get("is_admin"):
        return redirect(url_for("admin_login"))
    registrations = Registration.query.all()
    latest_data = {
        "name": session.get('name'),
        "age": session.get('age'),
        "email": session.get('email'),
        "phone": session.get('phone'),
        "service": session.get('service'),
    }
    return render_template(
        'admin_registrations.html',
        registrations=registrations,
        latest=latest_data,
        datetime=datetime
    )


@app.route('/admin/analytics/api')
def admin_analytics_api():
    if not session.get("is_admin"):
        return jsonify({'error': 'Unauthorized'}), 401
    analytics = calculate_analytics()
    return jsonify(analytics)


@app.route('/admin/page_analytics/api')
def admin_page_analytics_api():
    """API endpoint for live page visit analytics."""
    if not session.get("is_admin"):
        return jsonify({'error': 'Unauthorized'}), 401
    return jsonify(calculate_page_analytics())


@app.route('/admin/export_analytics')
def export_analytics():
    if not session.get("is_admin"):
        return redirect(url_for("admin_login"))
    analytics = calculate_analytics()
    analytics_file = os.path.join(UPLOAD_FOLDER, f"analytics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json")
    with open(analytics_file, 'w') as f:
        json.dump(analytics, f, indent=2)
    return send_file(analytics_file, as_attachment=True)


@app.route('/admin/download_responses')
def download_responses():
    if not session.get("is_admin"):
        return redirect(url_for("admin_login"))
    if not os.path.exists(EXCEL_FILE):
        return "No responses available to download.", 404
    return send_file(EXCEL_FILE, as_attachment=True, download_name="responses.xlsx")


@app.route('/admin/clear_excel', methods=['POST'])
def clear_excel():
    if not session.get("is_admin"):
        return redirect(url_for("admin_login"))
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        wb.save(EXCEL_FILE)
    return redirect(url_for('admin_panel'))


@app.route('/admin/delete_submission/<int:id>', methods=['POST'])
def delete_submission(id):
    try:
        submission = Registration.query.get(id)
        if submission:
            db.session.delete(submission)
            db.session.commit()
            return jsonify({"success": True, "message": "Submission deleted successfully."})
        return jsonify({"success": False, "message": "Submission not found."})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)})


@app.route('/admin/logout')
def admin_logout():
    session.pop("is_admin", None)
    return redirect(url_for("admin_login"))


# ── SPARK RUNNER GAME API ─────────────────────────────────────────────────────

@app.route('/api/save_progress', methods=['POST'])
def save_progress():
    try:
        data = request.get_json()
        if not data or 'glimmers' not in data:
            return jsonify({'success': False, 'message': 'Invalid data - glimmers count required'}), 400
        progress_data = load_game_data()
        new_entry = {
            'glimmers': data.get('glimmers', 0),
            'target': data.get('target', 10),
            'timestamp': data.get('timestamp', datetime.now().isoformat()),
            'session_id': len(progress_data) + 1,
            'user_agent': request.headers.get('User-Agent', 'Unknown'),
            'ip_address': request.remote_addr
        }
        progress_data.append(new_entry)
        save_game_data(progress_data)
        return jsonify({'success': True, 'message': 'Progress saved successfully', 'data': new_entry}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error saving progress: {str(e)}'}), 500


@app.route('/api/get_progress', methods=['GET'])
def get_progress():
    try:
        progress_data = load_game_data()
        return jsonify({'success': True, 'total_sessions': len(progress_data), 'data': progress_data}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error retrieving progress: {str(e)}'}), 500


@app.route('/api/get_stats', methods=['GET'])
def get_stats():
    try:
        progress_data = load_game_data()
        if not progress_data:
            return jsonify({'success': True, 'stats': {
                'total_sessions': 0, 'total_glimmers': 0, 'average_glimmers': 0,
                'highest_score': 0, 'latest_target': 10, 'last_played': None
            }}), 200
        total_glimmers = sum(entry.get('glimmers', 0) for entry in progress_data)
        highest_score = max(entry.get('glimmers', 0) for entry in progress_data)
        average_glimmers = total_glimmers / len(progress_data)
        stats = {
            'total_sessions': len(progress_data),
            'total_glimmers': total_glimmers,
            'average_glimmers': round(average_glimmers, 2),
            'highest_score': highest_score,
            'latest_target': progress_data[-1].get('target', 10),
            'last_played': progress_data[-1].get('timestamp')
        }
        return jsonify({'success': True, 'stats': stats}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error calculating stats: {str(e)}'}), 500


@app.route('/api/reset_progress', methods=['DELETE'])
def reset_progress():
    if not session.get("is_admin"):
        return jsonify({'success': False, 'message': 'Unauthorized - Admin access required'}), 401
    try:
        save_game_data([])
        return jsonify({'success': True, 'message': 'All game progress has been reset'}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error resetting progress: {str(e)}'}), 500


@app.route('/api/health', methods=['GET'])
def health_check():
    return jsonify({
        'status': 'healthy',
        'service': 'Mental Health Platform with Spark Runner',
        'timestamp': datetime.now().isoformat(),
        'version': '2.1'
    }), 200


# ── UTILITIES ─────────────────────────────────────────────────────────────────

def get_location(ip):
    try:
        if ip in ["127.0.0.1", "::1"] or ip.startswith(("192.168.", "10.", "172.")):
            return "Local Network"
        url = f"https://api.ipgeolocation.io/ipgeo?apiKey={IPGEO_API_KEY}&ip={ip}"
        response = requests.get(url, timeout=5)
        if response.status_code == 200:
            data = response.json()
            city = data.get("city", "Unknown")
            country = data.get("country_name", "Unknown")
            return f"{city}, {country}"
    except requests.exceptions.RequestException as e:
        print(f"Error fetching location: {e}")
    return "Location unavailable"


# ── RUN ───────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("=" * 60)
    print("Mental Health Platform with Spark Runner Game  v2.1")
    print("=" * 60)
    print("Server running on http://localhost:5000")
    print("\nPage visit tracking enabled on all routes.")
    print("Admin dashboard: /admin/login")
    print("=" * 60)
    app.run(debug=True)